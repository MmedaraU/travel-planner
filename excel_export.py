import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import database as db
from datetime import datetime
from currency import get_currency_symbol  # import from currency module


def export_profile_to_excel(exec_id, currency_symbol="$"):
    """Generate an Excel file with the executive's profile and memberships."""
    profile_data = db.get_full_executive_profile(exec_id)
    if not profile_data:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Profile"

    # --- Headers ---
    headers = ["Field", "Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # --- Profile Data ---
    row = 2
    fields = [
        ("Name", profile_data.get("Name", "")),
        ("Email", profile_data.get("Email", "")),
        ("Company", profile_data.get("Company", "")),
        ("Timezone", profile_data.get("Timezone", "")),
        ("Seat Preference", profile_data.get("Seat Preference", "")),
        ("Preferred Airline", profile_data.get("Preferred Airline", "")),
        ("Passport Number", profile_data.get("Passport Number", "")),
        ("TSA PreCheck", profile_data.get("TSA PreCheck", "")),
        ("Meal Preference", profile_data.get("Meal Preference", "")),
        ("Hotel Loyalty", profile_data.get("Hotel Loyalty", "")),
        ("Frequent Flyer #", profile_data.get("Frequent Flyer", "")),
        ("Dietary Restrictions", profile_data.get("Dietary", "")),
        ("Cost Center", profile_data.get("Cost Center", "")),
        ("Policy Notes", profile_data.get("Policy Notes", "")),
    ]

    for label, value in fields:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # --- Memberships Section (separate table) ---
    memberships = db.get_memberships(exec_id)
    if memberships:
        row += 1
        ws.cell(row=row, column=1, value="✈️ Memberships").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Category")
        ws.cell(row=row, column=2, value="Program")
        ws.cell(row=row, column=3, value="Number")
        for cell in ws[row]:
            cell.font = Font(bold=True)
        row += 1
        for m in memberships:
            ws.cell(row=row, column=1, value=m["category"].title())
            ws.cell(row=row, column=2, value=m["program_name"])
            ws.cell(row=row, column=3, value=m["membership_number"])
            row += 1

    # --- Auto-size columns ---
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_itinerary_to_excel(
    trip_id,
    display_currency_symbol="$",
    display_currency_code="USD",
    base_currency="USD",
):
    """
    Generate an Excel file with trip details, stops, and itinerary items.
    Shows amounts in the trip's Display Currency (symbol only, no conversion).
    """
    trip_data = db.get_trip(trip_id)
    if not trip_data:
        return None

    stops = db.get_trip_stops(trip_id)
    items = db.get_items_for_trip(trip_id)

    wb = Workbook()
    display_symbol = display_currency_symbol
    base_symbol = get_currency_symbol(base_currency)

    # --- Sheet 1: Trip Summary ---
    ws = wb.active
    ws.title = "Trip Summary"

    summary_data = [
        ("Trip Purpose", trip_data.get("purpose", "")),
        ("Destination", trip_data.get("destination", "")),
        ("Status", trip_data.get("status", "").title()),
        ("Budget", f"{display_symbol}{trip_data.get('budget', 0):,.2f}"),
        ("Display Currency", display_currency_code),
        ("Base Currency (Reporting)", base_currency),
        ("Departure City", trip_data.get("departure_city", "")),
        ("Departure Region", trip_data.get("departure_region", "")),
        ("Departure Country", trip_data.get("departure_country", "")),
        ("Start Date", trip_data.get("start_date", "")),
        ("End Date", trip_data.get("end_date", "")),
    ]

    for idx, (label, value) in enumerate(summary_data, 1):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)

    # --- Sheet 2: Stops ---
    ws_stops = wb.create_sheet("Stops")
    stop_headers = [
        "Order",
        "City",
        "Region",
        "Country",
        "Start Date",
        "End Date",
        "Notes",
    ]
    for col, header in enumerate(stop_headers, 1):
        cell = ws_stops.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

    for idx, stop in enumerate(stops, 2):
        ws_stops.cell(row=idx, column=1, value=stop.get("stop_order", ""))
        ws_stops.cell(row=idx, column=2, value=stop.get("city", ""))
        ws_stops.cell(row=idx, column=3, value=stop.get("region", ""))
        ws_stops.cell(row=idx, column=4, value=stop.get("country", ""))
        ws_stops.cell(row=idx, column=5, value=stop.get("start_date", ""))
        ws_stops.cell(row=idx, column=6, value=stop.get("end_date", ""))
        ws_stops.cell(row=idx, column=7, value=stop.get("notes", ""))

    # --- Sheet 3: Itinerary Items (with Display Currency) ---
    ws_items = wb.create_sheet("Itinerary Items")
    item_headers = [
        "Type",
        "Description",
        "Start Time",
        "End Time",
        "Location",
        "Original Amount",
        "Original Currency",
        f"Display Amount ({display_currency_code})",
        "Confirmed",
        "Confirmation Code",
        "Notes",
        "Receipt",
    ]
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

    for idx, item in enumerate(items, 2):
        orig_cost = item.get("cost", 0)
        orig_currency = item.get("cost_currency", "USD")
        # The Display Amount is just the original amount with the display symbol
        display_amount = f"{display_symbol}{orig_cost:.2f}" if orig_cost else ""

        ws_items.cell(row=idx, column=1, value=item.get("item_type", ""))
        ws_items.cell(row=idx, column=2, value=item.get("description", ""))
        ws_items.cell(row=idx, column=3, value=item.get("datetime_start", ""))
        ws_items.cell(row=idx, column=4, value=item.get("datetime_end", ""))
        ws_items.cell(row=idx, column=5, value=item.get("location", ""))
        ws_items.cell(
            row=idx,
            column=6,
            value=f"{orig_cost:.2f}" if orig_cost else "",
        )
        ws_items.cell(row=idx, column=7, value=orig_currency)
        ws_items.cell(row=idx, column=8, value=display_amount)
        ws_items.cell(
            row=idx, column=9, value="Yes" if item.get("is_confirmed") else "No"
        )
        ws_items.cell(row=idx, column=10, value=item.get("confirmation_code", ""))
        ws_items.cell(row=idx, column=11, value=item.get("notes", ""))
        ws_items.cell(row=idx, column=12, value=item.get("receipt_path", ""))

    # --- Auto-size all sheets ---
    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_spending_to_excel(
    summary_data,
    display_currency_symbol="$",
    display_currency_code="USD",
    base_currency="USD",
):
    """Generate an Excel file from the spending dashboard data."""
    if not summary_data:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Spending Dashboard"
    base_symbol = get_currency_symbol(base_currency)

    headers = [
        "Executive",
        "Company",
        "Destination",
        "Budget",
        "Total Spent",
        "Confirmed",
        "Estimated",
        "Status",
        "Base Currency",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    for idx, trip in enumerate(summary_data, 2):
        trip_base = trip.get("base_currency", "USD")
        trip_symbol = get_currency_symbol(trip_base)
        ws.cell(row=idx, column=1, value=trip.get("executive_name", ""))
        ws.cell(row=idx, column=2, value=trip.get("company_name", ""))
        ws.cell(row=idx, column=3, value=trip.get("destination", ""))
        ws.cell(row=idx, column=4, value=f"{trip_symbol}{trip.get('budget', 0):,.2f}")
        ws.cell(
            row=idx,
            column=5,
            value=f"{trip_symbol}{trip.get('total_spent', 0):,.2f}",
        )
        ws.cell(
            row=idx,
            column=6,
            value=f"{trip_symbol}{trip.get('confirmed_spent', 0):,.2f}",
        )
        ws.cell(
            row=idx,
            column=7,
            value=f"{trip_symbol}{trip.get('estimated_spent', 0):,.2f}",
        )
        ws.cell(row=idx, column=8, value=trip.get("status", "").title())
        ws.cell(row=idx, column=9, value=trip_base)

    if summary_data:
        total_row = len(summary_data) + 2
        ws.cell(row=total_row, column=3, value="TOTALS").font = Font(bold=True)
        ws.cell(
            row=total_row,
            column=4,
            value=f"{base_symbol}{sum(t['budget'] for t in summary_data):,.2f}",
        ).font = Font(bold=True)
        ws.cell(
            row=total_row,
            column=5,
            value=f"{base_symbol}{sum(t['total_spent'] for t in summary_data):,.2f}",
        ).font = Font(bold=True)
        ws.cell(
            row=total_row,
            column=6,
            value=f"{base_symbol}{sum(t['confirmed_spent'] for t in summary_data):,.2f}",
        ).font = Font(bold=True)
        ws.cell(
            row=total_row,
            column=7,
            value=f"{base_symbol}{sum(t['estimated_spent'] for t in summary_data):,.2f}",
        ).font = Font(bold=True)

    # Add a note about base currency
    note_row = len(summary_data) + 4 if summary_data else 1
    ws.cell(
        row=note_row, column=1, value=f"Base Currency for totals: {base_currency}"
    ).font = Font(italic=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================================================
# EXPENSE REPORT TO EXCEL (with full currency conversion)
# =========================================================
def export_expense_to_excel(
    items, trip_data, display_currency_symbol="$", base_currency="USD"
):
    """
    Generate an Excel file with the expense report structure:
    - Grouped by day
    - Time, Description, Type, Original Cost, Converted Cost, Receipt
    - Day subtotals
    - Grand totals (Confirmed, Estimated, Total)
    All costs are converted to base_currency using snapshot rates.
    """
    if not items:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    base_symbol = get_currency_symbol(base_currency)

    # --- Helper to add a formatted row ---
    def add_row(row_num, values, bold=False, bg_color=None):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if bold:
                cell.font = Font(bold=True)
            if bg_color:
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )

    # --- Group items by day ---
    sorted_items = sorted(
        items, key=lambda x: datetime.fromisoformat(x["datetime_start"])
    )
    days = {}
    for item in sorted_items:
        date_key = datetime.fromisoformat(item["datetime_start"]).strftime("%Y-%m-%d")
        if date_key not in days:
            days[date_key] = []
        days[date_key].append(item)

    # --- Headers ---
    headers = [
        "Date",
        "Time",
        "Description",
        "Type",
        "Original Amount",
        "Currency",
        f"Converted Amount ({base_currency})",
        "Receipt",
    ]
    add_row(1, headers, bold=True, bg_color="D3D3D3")

    row_num = 2
    grand_total = 0
    confirmed_total = 0
    estimated_total = 0

    for date_key, day_items in days.items():
        # --- Date Heading Row (merged) ---
        ws.cell(
            row=row_num,
            column=1,
            value=f"📅 {datetime.strptime(date_key, '%Y-%m-%d').strftime('%d-%m-%Y')}",
        ).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        row_num += 1

        day_total = 0
        for item in day_items:
            dt = datetime.fromisoformat(item["datetime_start"])
            time_str = dt.strftime("%H:%M")
            orig_cost = item.get("cost", 0)
            orig_currency = item.get("cost_currency", "USD")
            snapshot_rate = item.get("exchange_rate_snapshot", 1.0)
            converted_cost = orig_cost * snapshot_rate
            day_total += converted_cost
            grand_total += converted_cost
            if item.get("is_confirmed"):
                confirmed_total += converted_cost
            else:
                estimated_total += converted_cost

            receipt_filename = (
                os.path.basename(item.get("receipt_path", ""))
                if item.get("receipt_path")
                else "—"
            )

            add_row(
                row_num,
                [
                    dt.strftime("%d-%m-%Y"),
                    time_str,
                    item.get("description", ""),
                    item.get("item_type", ""),
                    f"{orig_cost:.2f}" if orig_cost else "",
                    orig_currency,
                    f"{base_symbol}{converted_cost:.2f}" if converted_cost else "",
                    receipt_filename,
                ],
            )
            row_num += 1

        # --- Day Total Row ---
        add_row(
            row_num,
            ["", "", "", "DAY TOTAL", "", "", f"{base_symbol}{day_total:.2f}", ""],
            bold=True,
            bg_color="E6E6E6",
        )
        row_num += 1

    # --- Summary Totals Section ---
    row_num += 1
    ws.cell(row=row_num, column=1, value="SUMMARY TOTALS").font = Font(
        bold=True, size=12
    )
    row_num += 1

    add_row(
        row_num,
        ["Confirmed (Booked)", f"{base_symbol}{confirmed_total:.2f}"],
        bold=True,
    )
    row_num += 1
    add_row(
        row_num,
        ["Estimated (Quoted)", f"{base_symbol}{estimated_total:.2f}"],
        bold=True,
    )
    row_num += 1
    add_row(
        row_num,
        ["GRAND TOTAL", f"{base_symbol}{grand_total:.2f}"],
        bold=True,
        bg_color="FFD700",
    )
    row_num += 1

    # --- Footnote about conversion ---
    row_num += 1
    ws.cell(
        row=row_num,
        column=1,
        value=f"Note: All amounts converted to {base_currency} using the exchange rate at the time of each expense (snapshot).",
    ).font = Font(italic=True)

    # --- Auto-size columns ---
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
